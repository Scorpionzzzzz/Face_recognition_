import os
os.environ['KMP_DUPLICATE_LIB_OK']='TRUE'
import sys
import cv2
import numpy as np
import pickle
import glob
import json
from datetime import datetime, timedelta
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QGridLayout, QLabel, QPushButton, 
                             QTableWidget, QTableWidgetItem, QTabWidget,
                             QFrame, QSplitter, QTextEdit, QComboBox,
                             QSpinBox, QCheckBox, QGroupBox, QMessageBox,
                             QFileDialog, QProgressBar, QSlider)
from PyQt5.QtCore import QTimer, QThread, pyqtSignal, Qt, QDateTime
from PyQt5.QtGui import QPixmap, QImage, QFont, QPalette, QColor, QIcon
from insightface.app import FaceAnalysis

class FaceRecognitionThread(QThread):
    """Thread riêng cho face recognition để không block GUI"""
    frame_ready = pyqtSignal(np.ndarray)
    face_detected = pyqtSignal(str, float, tuple)
    no_face = pyqtSignal()
    
    def __init__(self, app, database, threshold=0.4):
        super().__init__()
        self.app = app
        self.database = database
        self.threshold = threshold
        self.running = True
        self.cap = None
        
    def run(self):
        try:
            self.cap = cv2.VideoCapture(0)
            if not self.cap.isOpened():
                print("❌ Không thể mở camera")
                return
            
            # Cài đặt camera properties
            self.cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
            self.cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
            self.cap.set(cv2.CAP_PROP_FPS, 30)
            self.cap.set(cv2.CAP_PROP_AUTOFOCUS, 1)
            
            # Đợi camera khởi tạo
            import time
            time.sleep(1)
            
            # Đọc frame đầu tiên để kiểm tra
            ret, test_frame = self.cap.read()
            if not ret:
                print("❌ Camera không thể đọc frame")
                return
                
            print("✅ Camera đã sẵn sàng")
            
            while self.running:
                ret, frame = self.cap.read()
                if not ret:
                    print("⚠️ Không thể đọc frame từ camera")
                    continue
                    
                # Detect faces
                faces = self.app.get(frame)
                
                if len(faces) > 0:
                    # Xử lý tất cả khuôn mặt được phát hiện
                    best_face = None
                    best_score = -1
                    best_name = "Unknown"
                    
                    for face in faces:
                        x1, y1, x2, y2 = face.bbox.astype(int)
                        
                        # L2-normalize embedding trước khi nhận diện
                        emb_normalized = face.embedding / (np.linalg.norm(face.embedding) + 1e-12)
                        name, score = self.recognize_face(emb_normalized)
                        
                        # Vẽ bounding box cho tất cả khuôn mặt
                        if name == "Unknown":
                            cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 0, 255), 2)
                            cv2.putText(frame, "Unknown", 
                                       (x1, y1-10), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 255), 2)
                        else:
                            cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 2)
                            cv2.putText(frame, name, 
                                       (x1, y1-10), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)
                        
                        # Tìm khuôn mặt có điểm số cao nhất
                        if score > best_score:
                            best_score = score
                            best_name = name
                            best_face = (x1, y1, x2, y2)
                    
                    # Gửi thông tin khuôn mặt tốt nhất
                    if best_face:
                        self.face_detected.emit(best_name, best_score, best_face)
                else:
                    self.no_face.emit()
                
                self.frame_ready.emit(frame)
        except Exception as e:
            print(f"❌ Lỗi trong face recognition thread: {e}")
        finally:
            if self.cap:
                self.cap.release()
    
    def recognize_face(self, emb):
        # Embedding đã được L2-normalize từ FaceRecognitionThread
        best_match, best_score = "Unknown", -1
        for name, db_emb in self.database.items():
            sim = np.dot(emb, db_emb)
            if sim > best_score:
                best_match, best_score = name, sim
        
        if best_score < self.threshold:
            return "Unknown", best_score
        return best_match, best_score
    
    def stop(self):
        self.running = False
        if self.cap:
            self.cap.release()

class AttendanceSystem(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Hệ Thống Điểm Danh Tự Động - Face Recognition")
        self.setGeometry(100, 100, 1400, 900)
        
        # Khởi tạo ArcFace
        self.init_arcface()
        
        # Khởi tạo database
        self.load_database()
        
        # Tạo thư mục lưu CSV trước
        self.attendance_csv_dir = "attendance_csv"
        if not os.path.exists(self.attendance_csv_dir):
            os.makedirs(self.attendance_csv_dir)
        
        # Khởi tạo attendance data
        self.attendance_data = []
        self.load_attendance_data()
        
        # Thêm cơ chế chống spam thông báo
        self.last_attendance_time = {}  # Lưu thời gian điểm danh cuối cùng của mỗi người
        self.notification_cooldown = 5  # Thời gian chờ giữa các thông báo (giây)
        self.last_notification_time = {}  # Lưu thời gian thông báo cuối cùng
        
        # Khởi tạo UI
        self.init_ui()
        
        # Khởi tạo face recognition thread
        self.face_thread = None
        # Không tự động bắt đầu face recognition, để người dùng chọn
        
        # Đồng bộ hóa dữ liệu với CSV
        self.sync_data_with_csv()
    
    def init_arcface(self):
        """Khởi tạo ArcFace model"""
        try:
            self.app = FaceAnalysis(name="buffalo_l")
            self.app.prepare(ctx_id=0, det_size=(640, 640))
            print("✅ ArcFace model loaded successfully")
        except Exception as e:
            print(f"❌ Error loading ArcFace: {e}")
            QMessageBox.critical(self, "Lỗi", f"Không thể tải ArcFace model: {e}")
    
    def load_database(self):
        """Load face database từ evaluation results (đã được L2-normalize)"""
        try:
            eval_dirs = glob.glob("evaluation_results/evaluation_*")
            if not eval_dirs:
                raise FileNotFoundError("Không tìm thấy thư mục evaluation_results!")
            
            latest_eval_dir = max(eval_dirs)
            database_path = os.path.join(latest_eval_dir, "face_database.pkl")
            
            with open(database_path, "rb") as f:
                self.database = pickle.load(f)
            
            print(f"✅ Loaded database: {len(self.database)} people (L2-normalized)")
        except Exception as e:
            print(f"❌ Error loading database: {e}")
            self.database = {}
    
    def load_attendance_data(self):
        """Load attendance data từ các file CSV theo ngày"""
        try:
            self.attendance_data = []
            
            # Quét tất cả file CSV trong thư mục attendance_csv
            if os.path.exists(self.attendance_csv_dir):
                csv_files = [f for f in os.listdir(self.attendance_csv_dir) if f.endswith('.csv')]
                
                for csv_file in csv_files:
                    csv_path = os.path.join(self.attendance_csv_dir, csv_file)
                    
                    # Đọc file CSV
                    import pandas as pd
                    df = pd.read_csv(csv_path, sep='\t', encoding='utf-8-sig')
                    
                    # Lấy ngày từ tên file (attendance_YYYY-MM-DD.csv)
                    date_str = csv_file.replace('attendance_', '').replace('.csv', '')
                    
                    # Chuyển đổi dữ liệu từ CSV sang format attendance_data
                    for _, row in df.iterrows():
                        # Chỉ load những record đã điểm danh (True) và có thời gian
                        if row['Điểm Danh'] == 'True' and pd.notna(row['Thời Gian']) and row['Thời Gian'].strip() != '':
                            # Tạo datetime từ ngày và thời gian
                            time_str = row['Thời Gian']
                            datetime_str = f"{date_str}T{time_str}"
                            
                            attendance_record = {
                                'datetime': datetime_str,
                                'name': row['ID/Tên'],
                                'status': 'Present',
                                'score': 8.0,  # Điểm số mặc định
                                'action': 'Auto'
                            }
                            self.attendance_data.append(attendance_record)
            
            print(f"✅ Đã load {len(self.attendance_data)} records từ CSV files")
        except Exception as e:
            print(f"❌ Error loading attendance data từ CSV: {e}")
            self.attendance_data = []
    
    def save_attendance_data(self):
        """Lưu attendance data vào CSV theo ngày"""
        try:
            # Lưu CSV theo ngày
            self.save_daily_csv()
            print("✅ Đã lưu attendance data vào CSV")
        except Exception as e:
            print(f"❌ Error saving attendance data: {e}")
    
    def save_daily_csv(self):
        """Lưu attendance data theo ngày vào file CSV riêng biệt với cấu trúc mới"""
        try:
            # Nhóm dữ liệu theo ngày
            daily_data = {}
            for record in self.attendance_data:
                date = datetime.fromisoformat(record['datetime']).date()
                date_str = date.strftime('%Y-%m-%d')
                
                if date_str not in daily_data:
                    daily_data[date_str] = []
                daily_data[date_str].append(record)
            
            # Lưu từng ngày vào file CSV riêng
            for date_str, records in daily_data.items():
                csv_filename = os.path.join(self.attendance_csv_dir, f"attendance_{date_str}.csv")
                
                # Tạo danh sách tất cả người trong database cho ngày này
                all_people = list(self.database.keys())
                
                # Tạo DataFrame với cấu trúc mới
                csv_data = []
                for person in all_people:
                    # Kiểm tra xem người này đã điểm danh chưa
                    person_attendance = [r for r in records if r['name'] == person]
                    
                    if person_attendance:
                        # Đã điểm danh
                        attendance_time = datetime.fromisoformat(person_attendance[0]['datetime']).strftime('%H:%M:%S')
                        csv_data.append({
                            'Thời Gian': attendance_time,
                            'ID/Tên': person,
                            'Điểm Danh': 'True'
                        })
                    else:
                        # Chưa điểm danh
                        csv_data.append({
                            'Thời Gian': '',
                            'ID/Tên': person,
                            'Điểm Danh': 'False'
                        })
                
                # Sắp xếp theo tên
                csv_data.sort(key=lambda x: x['ID/Tên'])
                
                # Tạo DataFrame và lưu CSV
                import pandas as pd
                df = pd.DataFrame(csv_data)
                
                df.to_csv(csv_filename, index=False, encoding='utf-8-sig', sep='\t')
                print(f"✅ Đã lưu CSV cho ngày {date_str}: {csv_filename}")
            
            # Tạo CSV cho ngày hôm nay nếu chưa có
            today_str = datetime.now().strftime('%Y-%m-%d')
            today_csv_filename = os.path.join(self.attendance_csv_dir, f"attendance_{today_str}.csv")
            
            if not os.path.exists(today_csv_filename):
                # Tạo CSV cho ngày hôm nay với tất cả người chưa điểm danh
                all_people = list(self.database.keys())
                csv_data = []
                for person in all_people:
                    csv_data.append({
                        'Thời Gian': '',
                        'ID/Tên': person,
                        'Điểm Danh': 'False'
                    })
                
                # Sắp xếp theo tên
                csv_data.sort(key=lambda x: x['ID/Tên'])
                
                # Tạo DataFrame và lưu CSV
                import pandas as pd
                df = pd.DataFrame(csv_data)
                df.to_csv(today_csv_filename, index=False, encoding='utf-8-sig', sep='\t')
                print(f"✅ Đã tạo CSV cho ngày hôm nay: {today_csv_filename}")
                
        except Exception as e:
            print(f"❌ Error saving daily CSV: {e}")
    
    def init_ui(self):
        """Khởi tạo giao diện người dùng"""
        # Set style
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f0f0f0;
            }
            QTabWidget::pane {
                border: 1px solid #c0c0c0;
                background-color: white;
            }
            QTabBar::tab {
                background-color: #e0e0e0;
                padding: 8px 16px;
                margin-right: 2px;
            }
            QTabBar::tab:selected {
                background-color: #4a90e2;
                color: white;
            }
            QPushButton {
                background-color: #4a90e2;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #357abd;
            }
            QPushButton:pressed {
                background-color: #2d5aa0;
            }
            QTableWidget {
                gridline-color: #d0d0d0;
                selection-background-color: #4a90e2;
            }
            QHeaderView::section {
                background-color: #f8f9fa;
                padding: 4px;
                border: 1px solid #d0d0d0;
                font-weight: bold;
            }
        """)
        
        # Central widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # Main layout
        main_layout = QVBoxLayout(central_widget)
        
        # Header
        header = self.create_header()
        main_layout.addWidget(header)
        
        # Tab widget
        self.tab_widget = QTabWidget()
        main_layout.addWidget(self.tab_widget)
        
        # Tabs
        self.tab_widget.addTab(self.create_realtime_tab(), "🎥 Real-time Recognition")
        self.tab_widget.addTab(self.create_attendance_tab(), "📊 Lịch Sử Điểm Danh")
        self.tab_widget.addTab(self.create_statistics_tab(), "📈 Thống Kê")
        self.tab_widget.addTab(self.create_management_tab(), "⚙️ Quản Lý")
        
        # Status bar
        self.statusBar().showMessage("Hệ thống sẵn sàng")
    
    def create_header(self):
        """Tạo header cho ứng dụng"""
        header_frame = QFrame()
        header_frame.setFrameStyle(QFrame.Box)
        header_frame.setStyleSheet("QFrame { background-color: #2c3e50; border-radius: 8px; }")
        
        layout = QHBoxLayout(header_frame)
        
        # Logo/Title
        title_label = QLabel("HỆ THỐNG ĐIỂM DANH")
        title_label.setStyleSheet("""
            QLabel {
                color: white;
                font-size: 24px;
                font-weight: bold;
                padding: 10px;
            }
        """)
        layout.addWidget(title_label)
        
        # Current time
        self.time_label = QLabel()
        self.time_label.setStyleSheet("""
            QLabel {
                color: white;
                font-size: 16px;
                padding: 10px;
            }
        """)
        layout.addWidget(self.time_label)
        
        # Timer để cập nhật thời gian
        self.time_timer = QTimer()
        self.time_timer.timeout.connect(self.update_time)
        self.time_timer.start(1000)
        
        layout.addStretch()
        
        return header_frame
    
    def create_realtime_tab(self):
        """Tạo tab real-time recognition"""
        widget = QWidget()
        layout = QHBoxLayout(widget)
        
        # Left side - Camera feed
        left_frame = QFrame()
        left_frame.setFrameStyle(QFrame.Box)
        left_layout = QVBoxLayout(left_frame)
        
        # Camera label
        self.camera_label = QLabel()
        self.camera_label.setMinimumSize(640, 480)
        self.camera_label.setStyleSheet("""
            QLabel {
                border: 2px solid #c0c0c0;
                background-color: #2c3e50;
                color: white;
                font-size: 16px;
                font-weight: bold;
            }
        """)
        self.camera_label.setText("📷 Camera\nChưa khởi động")
        self.camera_label.setAlignment(Qt.AlignCenter)
        left_layout.addWidget(self.camera_label)
        
        # Control buttons
        control_layout = QHBoxLayout()
        
        self.start_btn = QPushButton("▶️ Bắt Đầu")
        self.start_btn.clicked.connect(self.start_face_recognition)
        control_layout.addWidget(self.start_btn)
        
        self.stop_btn = QPushButton("⏹️ Dừng")
        self.stop_btn.clicked.connect(self.stop_recognition)
        self.stop_btn.setEnabled(False)
        control_layout.addWidget(self.stop_btn)
        
        left_layout.addLayout(control_layout)
        
        # Right side - Recognition info
        right_frame = QFrame()
        right_frame.setFrameStyle(QFrame.Box)
        right_layout = QVBoxLayout(right_frame)
        
        # Current recognition
        current_group = QGroupBox("👤 Nhận Diện Hiện Tại")
        current_layout = QVBoxLayout(current_group)
        
        self.current_name_label = QLabel("Chưa có khuôn mặt")
        self.current_name_label.setStyleSheet("font-size: 18px; font-weight: bold; color: #666;")
        current_layout.addWidget(self.current_name_label)
        
        right_layout.addWidget(current_group)
        

        
        # Status display
        status_group = QGroupBox("📊 Trạng Thái")
        status_layout = QVBoxLayout(status_group)
        
        self.attendance_status_label = QLabel("Chưa điểm danh")
        self.attendance_status_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #666;")
        status_layout.addWidget(self.attendance_status_label)
        
        right_layout.addWidget(status_group)
        
        right_layout.addStretch()
        
        # Add to main layout
        layout.addWidget(left_frame, 2)
        layout.addWidget(right_frame, 1)
        
        return widget
    
    def create_attendance_tab(self):
        """Tạo tab lịch sử điểm danh theo ngày"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Date selection controls
        date_frame = QFrame()
        date_frame.setFrameStyle(QFrame.Box)
        date_layout = QHBoxLayout(date_frame)
        
        date_layout.addWidget(QLabel("📅 Chọn ngày:"))
        self.date_picker = QComboBox()
        self.update_date_picker()
        self.date_picker.currentTextChanged.connect(self.on_date_changed)
        date_layout.addWidget(self.date_picker)
        
        # Simulate date button
        self.simulate_date_btn = QPushButton("🎭 Mô Phỏng Ngày Khác")
        self.simulate_date_btn.clicked.connect(self.simulate_different_date)
        date_layout.addWidget(self.simulate_date_btn)
        
        # Current date display
        self.current_date_label = QLabel()
        self.current_date_label.setStyleSheet("font-weight: bold; color: #4a90e2; font-size: 14px;")
        date_layout.addWidget(self.current_date_label)
        
        date_layout.addStretch()
        layout.addWidget(date_frame)
        
        # Daily summary
        summary_frame = QFrame()
        summary_frame.setFrameStyle(QFrame.Box)
        summary_layout = QHBoxLayout(summary_frame)
        
        # Total people in database
        total_group = QGroupBox("👥 Tổng Số Người")
        total_layout = QVBoxLayout(total_group)
        self.total_people_label = QLabel(str(len(self.database)))
        self.total_people_label.setStyleSheet("font-size: 24px; font-weight: bold; color: #4a90e2;")
        total_layout.addWidget(self.total_people_label)
        summary_layout.addWidget(total_group)
        
        # Present today
        present_group = QGroupBox("✅ Điểm Danh Hôm Nay")
        present_layout = QVBoxLayout(present_group)
        self.present_today_label = QLabel("0")
        self.present_today_label.setStyleSheet("font-size: 24px; font-weight: bold; color: #28a745;")
        present_layout.addWidget(self.present_today_label)
        summary_layout.addWidget(present_group)
        
        # Absent today
        absent_group = QGroupBox("❌ Vắng Hôm Nay")
        absent_layout = QVBoxLayout(absent_group)
        self.absent_today_label = QLabel("0")
        self.absent_today_label.setStyleSheet("font-size: 24px; font-weight: bold; color: #dc3545;")
        absent_layout.addWidget(self.absent_today_label)
        summary_layout.addWidget(absent_group)
        
        # Attendance rate
        rate_group = QGroupBox("📊 Tỷ Lệ Điểm Danh")
        rate_layout = QVBoxLayout(rate_group)
        self.attendance_rate_label = QLabel("0%")
        self.attendance_rate_label.setStyleSheet("font-size: 24px; font-weight: bold; color: #ffc107;")
        rate_layout.addWidget(self.attendance_rate_label)
        summary_layout.addWidget(rate_group)
        
        layout.addWidget(summary_frame)
        
        # Attendance table for selected date
        table_group = QGroupBox("📋 Danh Sách Điểm Danh Theo Ngày")
        table_layout = QVBoxLayout(table_group)
        
        self.attendance_table = QTableWidget()
        self.attendance_table.setColumnCount(3)
        self.attendance_table.setHorizontalHeaderLabels([
            "Thời Gian", "Tên", "Hành Động"
        ])
        
        # Set column widths
        self.attendance_table.setColumnWidth(0, 150)
        self.attendance_table.setColumnWidth(1, 200)
        self.attendance_table.setColumnWidth(2, 100)
        
        table_layout.addWidget(self.attendance_table)
        layout.addWidget(table_group)
        
        # Action buttons
        action_layout = QHBoxLayout()
        
        self.export_btn = QPushButton("📤 Xuất Excel")
        self.export_btn.clicked.connect(self.export_attendance)
        action_layout.addWidget(self.export_btn)
        
        self.clear_history_btn = QPushButton("🗑️ Xóa Lịch Sử")
        self.clear_history_btn.clicked.connect(self.clear_attendance_history)
        action_layout.addWidget(self.clear_history_btn)
        
        action_layout.addStretch()
        
        layout.addLayout(action_layout)
        
        # Initialize with current date
        self.update_current_date_display()
        self.update_attendance_table()
        self.update_daily_summary()
        
        return widget
    
    def create_statistics_tab(self):
        """Tạo tab thống kê điểm danh"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Date range selection
        range_frame = QFrame()
        range_frame.setFrameStyle(QFrame.Box)
        range_layout = QHBoxLayout(range_frame)
        
        range_layout.addWidget(QLabel("📅 Từ ngày:"))
        self.start_date_picker = QComboBox()
        self.update_date_picker_for_stats()
        range_layout.addWidget(self.start_date_picker)
        
        range_layout.addWidget(QLabel("📅 Đến ngày:"))
        self.end_date_picker = QComboBox()
        self.update_date_picker_for_stats()
        range_layout.addWidget(self.end_date_picker)
        
        self.generate_stats_btn = QPushButton("📊 Tạo Thống Kê")
        self.generate_stats_btn.clicked.connect(self.generate_statistics)
        range_layout.addWidget(self.generate_stats_btn)
        
        layout.addWidget(range_frame)
        
        # Statistics display
        stats_frame = QFrame()
        stats_frame.setFrameStyle(QFrame.Box)
        stats_layout = QVBoxLayout(stats_frame)
        
        # Overall statistics
        overall_group = QGroupBox("📊 Thống Kê Tổng Quan")
        overall_layout = QGridLayout(overall_group)
        
        overall_layout.addWidget(QLabel("Tổng số ngày:"), 0, 0)
        self.total_days_label = QLabel("0")
        self.total_days_label.setStyleSheet("font-weight: bold; color: #4a90e2;")
        overall_layout.addWidget(self.total_days_label, 0, 1)
        
        overall_layout.addWidget(QLabel("Tổng số lần điểm danh:"), 1, 0)
        self.total_attendance_label = QLabel("0")
        self.total_attendance_label.setStyleSheet("font-weight: bold; color: #28a745;")
        overall_layout.addWidget(self.total_attendance_label, 1, 1)
        
        overall_layout.addWidget(QLabel("Tỷ lệ điểm danh trung bình:"), 2, 0)
        self.avg_attendance_rate_label = QLabel("0%")
        self.avg_attendance_rate_label.setStyleSheet("font-weight: bold; color: #ffc107;")
        overall_layout.addWidget(self.avg_attendance_rate_label, 2, 1)
        
        stats_layout.addWidget(overall_group)
        
        # Daily breakdown
        daily_group = QGroupBox("📅 Chi Tiết Theo Ngày")
        daily_layout = QVBoxLayout(daily_group)
        
        self.daily_stats_table = QTableWidget()
        self.daily_stats_table.setColumnCount(5)
        self.daily_stats_table.setHorizontalHeaderLabels([
            "Ngày", "Tổng Người", "Điểm Danh", "Vắng", "Tỷ Lệ"
        ])
        
        # Set column widths
        self.daily_stats_table.setColumnWidth(0, 120)
        self.daily_stats_table.setColumnWidth(1, 100)
        self.daily_stats_table.setColumnWidth(2, 100)
        self.daily_stats_table.setColumnWidth(3, 100)
        self.daily_stats_table.setColumnWidth(4, 100)
        
        daily_layout.addWidget(self.daily_stats_table)
        daily_group.setLayout(daily_layout)
        
        stats_layout.addWidget(daily_group)
        
        # Person statistics
        person_group = QGroupBox("👥 Thống Kê Theo Người")
        person_layout = QVBoxLayout(person_group)
        
        self.person_stats_table = QTableWidget()
        self.person_stats_table.setColumnCount(4)
        self.person_stats_table.setHorizontalHeaderLabels([
            "Tên", "Số Lần Điểm Danh", "Tỷ Lệ", "Ngày Cuối"
        ])
        
        # Set column widths
        self.person_stats_table.setColumnWidth(0, 200)
        self.person_stats_table.setColumnWidth(1, 150)
        self.person_stats_table.setColumnWidth(2, 100)
        self.person_stats_table.setColumnWidth(3, 120)
        
        person_layout.addWidget(self.person_stats_table)
        person_group.setLayout(person_layout)
        
        stats_layout.addWidget(person_group)
        
        stats_frame.setLayout(stats_layout)
        layout.addWidget(stats_frame)
        
        # Export statistics
        export_layout = QHBoxLayout()
        self.export_stats_btn = QPushButton("📤 Xuất Thống Kê CSV")
        self.export_stats_btn.clicked.connect(self.export_statistics)
        export_layout.addWidget(self.export_stats_btn)
        export_layout.addStretch()
        layout.addLayout(export_layout)
        
        return widget
    
    def create_management_tab(self):
        """Tạo tab quản lý"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Database info
        db_group = QGroupBox("🗄️ Thông Tin Database")
        db_layout = QGridLayout(db_group)
        
        db_layout.addWidget(QLabel("Số người trong database:"), 0, 0)
        self.db_count_label = QLabel(str(len(self.database)))
        db_layout.addWidget(self.db_count_label, 0, 1)
        
        db_layout.addWidget(QLabel("Danh sách người:"), 1, 0)
        self.db_list_label = QLabel(", ".join(self.database.keys()))
        self.db_list_label.setWordWrap(True)
        db_layout.addWidget(self.db_list_label, 1, 1)
        
        layout.addWidget(db_group)
        
        # CSV Directory info
        csv_group = QGroupBox("📁 Thông Tin CSV")
        csv_layout = QGridLayout(csv_group)
        
        csv_layout.addWidget(QLabel("Thư mục lưu CSV:"), 0, 0)
        csv_dir_label = QLabel(self.attendance_csv_dir)
        csv_dir_label.setStyleSheet("font-weight: bold; color: #28a745;")
        csv_layout.addWidget(csv_dir_label, 0, 1)
        
        # Đếm số file CSV
        csv_files = [f for f in os.listdir(self.attendance_csv_dir) if f.endswith('.csv')]
        csv_layout.addWidget(QLabel("Số file CSV:"), 1, 0)
        csv_count_label = QLabel(str(len(csv_files)))
        csv_count_label.setStyleSheet("font-weight: bold; color: #4a90e2;")
        csv_layout.addWidget(csv_count_label, 1, 1)
        
        layout.addWidget(csv_group)
        
        # System status
        status_group = QGroupBox("📊 Trạng Thái Hệ Thống")
        status_layout = QGridLayout(status_group)
        
        status_layout.addWidget(QLabel("ArcFace Model:"), 0, 0)
        self.model_status_label = QLabel("✅ Đã tải")
        status_layout.addWidget(self.model_status_label, 0, 1)
        
        status_layout.addWidget(QLabel("Camera:"), 1, 0)
        self.camera_status_label = QLabel("⏸️ Đã dừng")
        status_layout.addWidget(self.camera_status_label, 1, 1)
        
        status_layout.addWidget(QLabel("Recognition Thread:"), 2, 0)
        self.thread_status_label = QLabel("⏸️ Đã dừng")
        status_layout.addWidget(self.thread_status_label, 2, 1)
        
        layout.addWidget(status_group)
        
        # Notification settings
        notification_group = QGroupBox("🔔 Cài Đặt Thông Báo")
        notification_layout = QGridLayout(notification_group)
        
        notification_layout.addWidget(QLabel("Thời gian chờ thông báo (giây):"), 0, 0)
        self.cooldown_spinbox = QSpinBox()
        self.cooldown_spinbox.setRange(1, 60)
        self.cooldown_spinbox.setValue(self.notification_cooldown)
        self.cooldown_spinbox.valueChanged.connect(self.update_notification_cooldown)
        notification_layout.addWidget(self.cooldown_spinbox, 0, 1)
        
        notification_layout.addWidget(QLabel("Trạng thái cooldown:"), 1, 0)
        self.cooldown_status_label = QLabel("✅ Hoạt động")
        self.cooldown_status_label.setStyleSheet("color: #28a745; font-weight: bold;")
        notification_layout.addWidget(self.cooldown_status_label, 1, 1)
        
        layout.addWidget(notification_group)
        
        # Actions
        action_group = QGroupBox("🔧 Thao Tác")
        action_layout = QVBoxLayout(action_group)
        
        self.reload_db_btn = QPushButton("🔄 Tải Lại Database")
        self.reload_db_btn.clicked.connect(self.reload_database)
        action_layout.addWidget(self.reload_db_btn)
        
        self.test_camera_btn = QPushButton("📷 Kiểm Tra Camera")
        self.test_camera_btn.clicked.connect(self.test_camera)
        action_layout.addWidget(self.test_camera_btn)
        
        self.test_notification_btn = QPushButton("🔔 Test Thông Báo")
        self.test_notification_btn.clicked.connect(self.test_notification)
        action_layout.addWidget(self.test_notification_btn)
        
        self.sync_csv_btn = QPushButton("🔄 Đồng Bộ CSV")
        self.sync_csv_btn.clicked.connect(self.sync_data_with_csv)
        action_layout.addWidget(self.sync_csv_btn)
        
        action_layout.addStretch()
        
        layout.addWidget(action_group)
        
        layout.addStretch()
        
        return widget
    
    def start_face_recognition(self):
        """Bắt đầu face recognition thread"""
        if self.face_thread is None or not self.face_thread.isRunning():
            self.face_thread = FaceRecognitionThread(self.app, self.database, threshold=0.4)
            self.face_thread.frame_ready.connect(self.update_camera_feed)
            self.face_thread.face_detected.connect(self.on_face_detected)
            self.face_thread.no_face.connect(self.on_no_face)
            self.face_thread.start()
            
            self.start_btn.setEnabled(False)
            self.stop_btn.setEnabled(True)
            self.camera_status_label.setText("✅ Đang hoạt động")
            self.thread_status_label.setText("✅ Đang chạy")
            self.statusBar().showMessage("Face recognition đã bắt đầu")
            
            # Cập nhật camera label
            self.camera_label.setText("📷 Camera\nĐang khởi động...")
            self.camera_label.setStyleSheet("""
                QLabel {
                    border: 2px solid #c0c0c0;
                    background-color: #2c3e50;
                    color: white;
                    font-size: 16px;
                    font-weight: bold;
                }
            """)
            
            # Reset trạng thái điểm danh
            self.attendance_status_label.setText("Chưa điểm danh")
            self.attendance_status_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #666;")
    
    def stop_recognition(self):
        """Dừng face recognition thread"""
        if self.face_thread and self.face_thread.isRunning():
            self.face_thread.stop()
            self.face_thread.wait()
            
            self.start_btn.setEnabled(True)
            self.stop_btn.setEnabled(False)
            self.camera_status_label.setText("⏸️ Đã dừng")
            self.thread_status_label.setText("⏸️ Đã dừng")
            self.statusBar().showMessage("Face recognition đã dừng")
            
            # Reset camera label
            self.camera_label.setText("📷 Camera\nChưa khởi động")
            self.camera_label.setStyleSheet("""
                QLabel {
                    border: 2px solid #c0c0c0;
                    background-color: #2c3e50;
                    color: white;
                    font-size: 16px;
                    font-weight: bold;
                }
            """)
            
            # Reset trạng thái điểm danh
            self.attendance_status_label.setText("Chưa điểm danh")
            self.attendance_status_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #666;")
    
    def update_camera_feed(self, frame):
        """Cập nhật camera feed"""
        try:
            if frame is None or frame.size == 0:
                return
                
            rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            h, w, ch = rgb_frame.shape
            
            if h <= 0 or w <= 0:
                return
                
            bytes_per_line = ch * w
            qt_image = QImage(rgb_frame.data, w, h, bytes_per_line, QImage.Format_RGB888)
            
            if qt_image.isNull():
                return
                
            pixmap = QPixmap.fromImage(qt_image)
            if not pixmap.isNull():
                scaled_pixmap = pixmap.scaled(
                    self.camera_label.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation
                )
                self.camera_label.setPixmap(scaled_pixmap)
        except Exception as e:
            print(f"❌ Lỗi cập nhật camera feed: {e}")
    
    def on_face_detected(self, name, score, bbox):
        """Xử lý khi phát hiện khuôn mặt"""
        self.current_name_label.setText(name)
        
        if name != "Unknown":
            self.current_name_label.setStyleSheet("font-size: 18px; font-weight: bold; color: #28a745;")
            
            # Kiểm tra cooldown trước khi hiển thị thông báo
            current_time = datetime.now()
            if name in self.last_notification_time:
                time_since_last = (current_time - self.last_notification_time[name]).total_seconds()
                if time_since_last < self.notification_cooldown:
                    # Đang trong cooldown, chỉ cập nhật UI mà không gửi thông báo
                    remaining_time = self.notification_cooldown - time_since_last
                    self.statusBar().showMessage(f"{name} - Cooldown: {remaining_time:.1f}s còn lại")
                    return
            
            # Tự động điểm danh
            self.auto_mark_attendance(name, score)
        else:
            self.current_name_label.setStyleSheet("font-size: 18px; font-weight: bold; color: #dc3545;")
            self.attendance_status_label.setText("Không nhận diện được")
            self.attendance_status_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #dc3545;")
    
    def on_no_face(self):
        """Xử lý khi không có khuôn mặt"""
        self.current_name_label.setText("Chưa có khuôn mặt")
        self.current_name_label.setStyleSheet("font-size: 18px; font-weight: bold; color: #666;")
        self.attendance_status_label.setText("Chưa điểm danh")
        self.attendance_status_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #666;")
    
    def auto_mark_attendance(self, name, score):
        """Tự động điểm danh khi nhận diện được khuôn mặt với cơ chế chống spam"""
        current_time = datetime.now()
        today = current_time.date()
        
        # Kiểm tra cooldown thông báo để tránh spam
        if name in self.last_notification_time:
            time_since_last_notification = (current_time - self.last_notification_time[name]).total_seconds()
            if time_since_last_notification < self.notification_cooldown:
                # Chưa đủ thời gian để hiển thị thông báo mới
                return
        
        # Kiểm tra xem đã điểm danh hôm nay chưa
        existing_attendance = [a for a in self.attendance_data 
                             if a['name'] == name and 
                             datetime.fromisoformat(a['datetime']).date() == today]
        
        if existing_attendance:
            # Đã điểm danh rồi - chỉ hiển thị thông báo nếu đủ thời gian cooldown
            self.attendance_status_label.setText(f"✅ {name} đã điểm danh hôm nay")
            self.attendance_status_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #28a745;")
            self.statusBar().showMessage(f"{name} đã được điểm danh hôm nay")
            
            # Cập nhật thời gian thông báo cuối cùng
            self.last_notification_time[name] = current_time
            
            # Hiển thị thông báo đẩy cho trường hợp đã điểm danh (chỉ khi đủ cooldown)
            self.show_push_notification(f"ℹ️ {name} đã điểm danh hôm nay!", 3000)
        else:
            # Chưa điểm danh, tiến hành điểm danh
            attendance_record = {
                'datetime': current_time.isoformat(),
                'name': name,
                'status': 'Present',
                'score': score,
                'action': 'Auto'
            }
            
            self.attendance_data.append(attendance_record)
            self.save_attendance_data()  # Sẽ tự động lưu cả JSON và CSV
            self.update_attendance_table()
            self.update_daily_summary()
            
            # Cập nhật thời gian điểm danh cuối cùng
            self.last_attendance_time[name] = current_time
            self.last_notification_time[name] = current_time
            
            # Cập nhật trạng thái
            self.attendance_status_label.setText(f"✅ Đã điểm danh {name} thành công!")
            self.attendance_status_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #28a745;")
            self.statusBar().showMessage(f"Đã điểm danh {name} - {current_time.strftime('%H:%M:%S')}")
            
            # Hiển thị thông báo đẩy tự động biến mất sau 3 giây
            self.show_push_notification(f"✅ Đã điểm danh {name} thành công!", 3000)
    
    def mark_attendance(self):
        """Điểm danh thủ công (giữ lại để tương thích)"""
        name = self.current_name_label.text()
        if name != "Unknown" and name != "Chưa có khuôn mặt":
            self.auto_mark_attendance(name, 0.0)  # Điểm số không hiển thị nên có thể để 0
    
    def clear_recognition(self):
        """Xóa kết quả nhận diện hiện tại"""
        self.current_name_label.setText("Chưa có khuôn mặt")
        self.current_name_label.setStyleSheet("font-size: 18px; font-weight: bold; color: #666;")
        self.attendance_status_label.setText("Chưa điểm danh")
        self.attendance_status_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #666;")
    
    def show_push_notification(self, message, duration=3000):
        """Hiển thị thông báo đẩy tự động biến mất"""
        # Xóa notification cũ nếu có
        if hasattr(self, 'notification_widget'):
            self.hide_notification()
        
        # Xác định màu sắc dựa trên nội dung message
        if "✅ Đã điểm danh" in message:
            bg_color = "#28a745"  # Xanh lá - thành công
            border_color = "#1e7e34"
        elif "ℹ️" in message:
            bg_color = "#17a2b8"  # Xanh dương - thông tin
            border_color = "#117a8b"
        else:
            bg_color = "#ffc107"  # Vàng - cảnh báo
            border_color = "#e0a800"
        
        # Tạo notification widget
        notification = QLabel(message)
        notification.setStyleSheet(f"""
            QLabel {{
                background-color: {bg_color};
                color: white;
                padding: 15px 20px;
                border-radius: 8px;
                font-size: 16px;
                font-weight: bold;
                border: 2px solid {border_color};
                min-height: 50px;
            }}
        """)
        notification.setAlignment(Qt.AlignCenter)
        notification.setWordWrap(True)
        
        # Tạo widget container
        notification_widget = QWidget()
        notification_widget.setStyleSheet("background-color: transparent;")
        
        # Tạo layout cho notification
        notification_layout = QVBoxLayout(notification_widget)
        notification_layout.setContentsMargins(0, 0, 0, 0)
        notification_layout.addWidget(notification)
        
        # Thêm vào main window
        self.notification_widget = notification_widget
        self.notification_widget.setParent(self)
        
        # Đặt vị trí (góc trên bên phải, trên header)
        self.notification_widget.setFixedSize(400, 100)
        
        # Tính toán vị trí chính xác
        x_pos = self.width() - 420
        y_pos = 10
        
        # Đảm bảo không vượt quá biên
        if x_pos < 0:
            x_pos = 10
        if y_pos < 0:
            y_pos = 10
            
        self.notification_widget.move(x_pos, y_pos)
        
        # Hiển thị notification
        self.notification_widget.show()
        self.notification_widget.raise_()
        self.notification_widget.activateWindow()
        
        # Force update
        self.notification_widget.repaint()
        
        # Timer để tự động ẩn notification
        QTimer.singleShot(duration, self.hide_notification)
        
        print(f"🔔 Hiển thị thông báo: {message} (sẽ ẩn sau {duration/1000}s)")
    
    def hide_notification(self):
        """Ẩn thông báo đẩy"""
        if hasattr(self, 'notification_widget'):
            self.notification_widget.hide()
            self.notification_widget.deleteLater()
            delattr(self, 'notification_widget')
            print("🔔 Đã ẩn thông báo")
    
    def sync_data_with_csv(self):
        """Đồng bộ hóa dữ liệu với CSV files"""
        try:
            print("🔄 Đang đồng bộ hóa dữ liệu với CSV...")
            
            # Đảm bảo có CSV cho ngày hôm nay
            self.save_daily_csv()
            
            # Reload attendance data từ CSV
            self.load_attendance_data()
            
            # Cập nhật UI
            self.update_date_picker()
            self.update_attendance_table()
            self.update_daily_summary()
            
            print(f"✅ Đã đồng bộ hóa: {len(self.attendance_data)} records")
            
            # Hiển thị thông báo thành công
            self.show_push_notification("🔄 Đã đồng bộ hóa dữ liệu với CSV!", 2000)
        except Exception as e:
            print(f"❌ Lỗi đồng bộ hóa: {e}")
            self.show_push_notification("❌ Lỗi đồng bộ hóa CSV!", 3000)
    
    # Method update_threshold đã được bỏ vì threshold cố định
    
    def update_time(self):
        """Cập nhật thời gian hiện tại và kiểm tra ngày mới để reset cooldown"""
        current_time = QDateTime.currentDateTime()
        self.time_label.setText(current_time.toString("dd/MM/yyyy hh:mm:ss"))
        
        # Kiểm tra nếu sang ngày mới thì reset cooldown
        current_date = current_time.date().toString("yyyy-MM-dd")
        if not hasattr(self, 'current_date_str'):
            self.current_date_str = current_date
        elif self.current_date_str != current_date:
            # Sang ngày mới, reset cooldown
            self.last_notification_time.clear()
            self.current_date_str = current_date
            print(f"🔄 Đã sang ngày mới ({current_date}), reset cooldown thông báo")
    
    def update_date_picker(self):
        """Cập nhật danh sách ngày trong date picker"""
        dates = set()
        for record in self.attendance_data:
            date = datetime.fromisoformat(record['datetime']).date()
            dates.add(date.strftime('%d/%m/%Y'))
        
        # Thêm ngày hôm nay nếu chưa có
        today = datetime.now().date().strftime('%d/%m/%Y')
        if today not in dates:
            dates.add(today)
        
        self.date_picker.clear()
        for date in sorted(dates, reverse=True):
            self.date_picker.addItem(date)
        
        # Chọn ngày hôm nay mặc định
        if today in dates:
            self.date_picker.setCurrentText(today)
    
    def update_date_picker_for_stats(self):
        """Cập nhật danh sách ngày cho thống kê"""
        dates = set()
        for record in self.attendance_data:
            date = datetime.fromisoformat(record['datetime']).date()
            dates.add(date.strftime('%d/%m/%Y'))
        
        # Thêm ngày hôm nay nếu chưa có
        today = datetime.now().date().strftime('%d/%m/%Y')
        if today not in dates:
            dates.add(today)
        
        sorted_dates = sorted(dates, reverse=True)
        
        # Kiểm tra xem các date picker đã được tạo chưa
        if hasattr(self, 'start_date_picker') and hasattr(self, 'end_date_picker'):
            self.start_date_picker.clear()
            self.end_date_picker.clear()
            
            for date in sorted_dates:
                self.start_date_picker.addItem(date)
                self.end_date_picker.addItem(date)
            
            # Chọn khoảng thời gian mặc định (7 ngày gần nhất)
            if len(sorted_dates) >= 7:
                self.start_date_picker.setCurrentText(sorted_dates[6])
                self.end_date_picker.setCurrentText(sorted_dates[0])
            else:
                self.start_date_picker.setCurrentText(sorted_dates[-1] if sorted_dates else today)
                self.end_date_picker.setCurrentText(sorted_dates[0] if sorted_dates else today)
    
    def update_current_date_display(self):
        """Cập nhật hiển thị ngày hiện tại"""
        if self.date_picker.count() > 0:
            current_date = self.date_picker.currentText()
            self.current_date_label.setText(f"Ngày hiện tại: {current_date}")
    
    def on_date_changed(self):
        """Xử lý khi thay đổi ngày"""
        self.update_current_date_display()
        self.update_attendance_table()
        self.update_daily_summary()
    
    def simulate_different_date(self):
        """Mô phỏng điểm danh cho ngày khác"""
        from PyQt5.QtWidgets import QInputDialog
        
        # Lấy danh sách ngày có sẵn
        available_dates = []
        for i in range(self.date_picker.count()):
            available_dates.append(self.date_picker.itemText(i))
        
        if not available_dates:
            QMessageBox.warning(self, "Cảnh báo", "Không có dữ liệu điểm danh nào!")
            return
        
        # Chọn ngày để mô phỏng
        date, ok = QInputDialog.getItem(
            self, "Mô Phỏng Điểm Danh", 
            "Chọn ngày để mô phỏng:", available_dates, 0, False
        )
        
        if ok and date:
            # Chọn người để mô phỏng điểm danh
            person, ok = QInputDialog.getItem(
                self, "Mô Phỏng Điểm Danh", 
                "Chọn người để điểm danh:", list(self.database.keys()), 0, False
            )
            
            if ok and person:
                # Tạo điểm danh mô phỏng
                target_date = datetime.strptime(date, '%d/%m/%Y').date()
                current_time = datetime.now()
                simulated_time = datetime.combine(target_date, current_time.time())
                
                # Kiểm tra xem đã điểm danh chưa
                existing_attendance = [a for a in self.attendance_data 
                                     if a['name'] == person and 
                                     datetime.fromisoformat(a['datetime']).date() == target_date]
                
                if existing_attendance:
                    QMessageBox.information(self, "Thông báo", f"{person} đã được điểm danh vào ngày {date}!")
                    return
                
                # Thêm điểm danh mô phỏng
                attendance_record = {
                    'datetime': simulated_time.isoformat(),
                    'name': person,
                    'status': 'Present',
                    'score': 8.5,  # Điểm số mô phỏng
                    'action': 'Simulated'
                }
                
                self.attendance_data.append(attendance_record)
                self.save_attendance_data()  # Sẽ tự động lưu cả JSON và CSV
                
                # Cập nhật UI
                self.update_date_picker()
                self.update_attendance_table()
                self.update_daily_summary()
                
                QMessageBox.information(self, "Thành công", f"Đã mô phỏng điểm danh {person} vào ngày {date}!\nDữ liệu đã được lưu vào CSV theo ngày.")
    
    def update_daily_summary(self):
        """Cập nhật thống kê ngày hiện tại"""
        if not hasattr(self, 'date_picker') or self.date_picker.count() == 0:
            return
            
        selected_date = self.date_picker.currentText()
        target_date = datetime.strptime(selected_date, '%d/%m/%Y').date()
        
        # Lọc dữ liệu theo ngày
        daily_attendance = [a for a in self.attendance_data 
                           if datetime.fromisoformat(a['datetime']).date() == target_date]
        
        # Cập nhật labels
        total_people = len(self.database)
        present_count = len(daily_attendance)
        absent_count = total_people - present_count
        attendance_rate = (present_count / total_people * 100) if total_people > 0 else 0
        
        if hasattr(self, 'total_people_label'):
            self.total_people_label.setText(str(total_people))
        if hasattr(self, 'present_today_label'):
            self.present_today_label.setText(str(present_count))
        if hasattr(self, 'absent_today_label'):
            self.absent_today_label.setText(str(absent_count))
        if hasattr(self, 'attendance_rate_label'):
            self.attendance_rate_label.setText(f"{attendance_rate:.1f}%")
    
    def apply_filters(self):
        """Áp dụng bộ lọc cho attendance table"""
        self.update_attendance_table()
      
    def update_attendance_table(self):
        """Cập nhật attendance table theo ngày được chọn"""
        if not hasattr(self, 'date_picker') or self.date_picker.count() == 0:
            return
            
        selected_date = self.date_picker.currentText()
        target_date = datetime.strptime(selected_date, '%d/%m/%Y').date()
        
        # Lọc dữ liệu theo ngày
        filtered_data = [r for r in self.attendance_data 
                         if datetime.fromisoformat(r['datetime']).date() == target_date]
        
        # Cập nhật table
        if hasattr(self, 'attendance_table'):
            self.attendance_table.setRowCount(len(filtered_data))
            
            for row, record in enumerate(filtered_data):
                dt = datetime.fromisoformat(record['datetime'])
                
                # Thời gian
                time_item = QTableWidgetItem(dt.strftime('%H:%M:%S'))
                self.attendance_table.setItem(row, 0, time_item)
                
                # Tên
                name_item = QTableWidgetItem(record['name'])
                self.attendance_table.setItem(row, 1, name_item)
                
                # Hành động
                action_item = QTableWidgetItem(record['action'])
                self.attendance_table.setItem(row, 2, action_item)
    
    def export_attendance(self):
        """Xuất attendance data ra Excel với các cột rõ ràng"""
        try:
            # Xuất Excel với cấu trúc rõ ràng
            filename, _ = QFileDialog.getSaveFileName(
                self, "Lưu file Excel", "", "Excel Files (*.xlsx)"
            )
            if filename:
                import pandas as pd
                
                # Tạo dữ liệu cho Excel với cấu trúc rõ ràng
                excel_data = []
                
                # Lấy tất cả người trong database
                all_people = list(self.database.keys())
                
                # Nhóm dữ liệu theo ngày
                daily_data = {}
                for record in self.attendance_data:
                    date = datetime.fromisoformat(record['datetime']).date()
                    date_str = date.strftime('%Y-%m-%d')
                    
                    if date_str not in daily_data:
                        daily_data[date_str] = []
                    daily_data[date_str].append(record)
                
                # Tạo dữ liệu cho Excel
                for date_str in sorted(daily_data.keys(), reverse=True):
                    records = daily_data[date_str]
                    
                    # Thêm header cho ngày
                    excel_data.append({
                        'Ngày': f"=== {date_str} ===",
                        'Thời Gian': '',
                        'ID/Tên': '',
                        'Điểm Danh': '',
                        'Ghi Chú': ''
                    })
                    
                    # Thêm dữ liệu cho từng người
                    for person in sorted(all_people):
                        person_attendance = [r for r in records if r['name'] == person]
                        
                        if person_attendance:
                            # Đã điểm danh
                            attendance_time = datetime.fromisoformat(person_attendance[0]['datetime']).strftime('%H:%M:%S')
                            excel_data.append({
                                'Ngày': date_str,
                                'Thời Gian': attendance_time,
                                'ID/Tên': person,
                                'Điểm Danh': '✅ CÓ',
                                'Ghi Chú': ''
                            })
                        else:
                            # Chưa điểm danh
                            excel_data.append({
                                'Ngày': date_str,
                                'Thời Gian': '',
                                'ID/Tên': person,
                                'Điểm Danh': '❌ VẮNG',
                                'Ghi Chú': ''
                            })
                    
                    # Thêm dòng trống giữa các ngày
                    excel_data.append({
                        'Ngày': '',
                        'Thời Gian': '',
                        'ID/Tên': '',
                        'Điểm Danh': '',
                        'Ghi Chú': ''
                    })
                
                # Tạo DataFrame và xuất Excel
                df = pd.DataFrame(excel_data)
                
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Điểm Danh', index=False)
                    
                    # Lấy worksheet để định dạng
                    worksheet = writer.sheets['Điểm Danh']
                    
                    # Định dạng header
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    
                    # Style cho header
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    header_alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Style cho ngày
                    date_font = Font(bold=True, color="FFFFFF", size=14)
                    date_fill = PatternFill(start_color="C05050", end_color="C05050", fill_type="solid")
                    date_alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Style cho có điểm danh
                    present_font = Font(bold=True, color="FFFFFF")
                    present_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
                    
                    # Style cho vắng
                    absent_font = Font(bold=True, color="FFFFFF")
                    absent_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    
                    # Border style
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # Áp dụng style cho header
                    for col in range(1, 6):  # A đến E
                        cell = worksheet.cell(row=1, column=col)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = thin_border
                    
                    # Áp dụng style cho các dòng dữ liệu
                    for row in range(2, len(excel_data) + 1):
                        for col in range(1, 6):
                            cell = worksheet.cell(row=row, column=col)
                            cell.border = thin_border
                            
                            # Style cho header ngày
                            if col == 1 and cell.value and '===' in str(cell.value):
                                cell.font = date_font
                                cell.fill = date_fill
                                cell.alignment = date_alignment
                            
                            # Style cho điểm danh
                            if col == 4:  # Cột Điểm Danh
                                if cell.value == '✅ CÓ':
                                    cell.font = present_font
                                    cell.fill = present_fill
                                    cell.alignment = Alignment(horizontal="center")
                                elif cell.value == '❌ VẮNG':
                                    cell.font = absent_font
                                    cell.fill = absent_fill
                                    cell.alignment = Alignment(horizontal="center")
                    
                    # Điều chỉnh độ rộng cột
                    worksheet.column_dimensions['A'].width = 15  # Ngày
                    worksheet.column_dimensions['B'].width = 12  # Thời Gian
                    worksheet.column_dimensions['C'].width = 25  # ID/Tên
                    worksheet.column_dimensions['D'].width = 15  # Điểm Danh
                    worksheet.column_dimensions['E'].width = 20  # Ghi Chú
                
                QMessageBox.information(self, "Thành công", f"Đã xuất dữ liệu ra Excel với định dạng đẹp: {filename}")
            
            # Vẫn xuất CSV theo ngày
            self.save_daily_csv()
            
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Không thể xuất Excel: {e}")
    
    def clear_attendance_history(self):
        """Xóa lịch sử điểm danh"""
        reply = QMessageBox.question(
            self, "Xác nhận", 
            "Bạn có chắc chắn muốn xóa toàn bộ lịch sử điểm danh?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            self.attendance_data = []
            self.save_attendance_data()
            self.update_attendance_table()
            self.update_date_picker()
            self.update_daily_summary()
            QMessageBox.information(self, "Thành công", "Đã xóa toàn bộ lịch sử điểm danh")
    
    def reload_database(self):
        """Tải lại database"""
        try:
            self.load_database()
            self.db_count_label.setText(str(len(self.database)))
            self.db_list_label.setText(", ".join(self.database.keys()))
            
            QMessageBox.information(self, "Thành công", "Đã tải lại database thành công")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Không thể tải lại database: {e}")
    
    def test_camera(self):
        """Kiểm tra camera"""
        try:
            cap = cv2.VideoCapture(0)
            if cap.isOpened():
                ret, frame = cap.read()
                cap.release()
                
                if ret:
                    QMessageBox.information(self, "Thành công", "Camera hoạt động bình thường")
                else:
                    QMessageBox.warning(self, "Cảnh báo", "Camera không thể đọc frame")
            else:
                QMessageBox.critical(self, "Lỗi", "Không thể mở camera")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Lỗi kiểm tra camera: {e}")
    
    def test_notification(self):
        """Test thông báo đẩy"""
        self.show_push_notification("🧪 Test thông báo đẩy - 3 giây!", 3000)
    
    def update_notification_cooldown(self, value):
        """Cập nhật thời gian cooldown thông báo"""
        self.notification_cooldown = value
        self.cooldown_status_label.setText(f"✅ {value}s cooldown")
        self.cooldown_status_label.setStyleSheet("color: #28a745; font-weight: bold;")
        print(f"🔔 Đã cập nhật cooldown thông báo: {value} giây")
    
    def closeEvent(self, event):
        """Xử lý khi đóng ứng dụng"""
        if self.face_thread and self.face_thread.isRunning():
            self.face_thread.stop()
            self.face_thread.wait()
        
        # Lưu attendance data vào CSV
        self.save_attendance_data()
        
        event.accept()
    
    def generate_statistics(self):
        """Tạo thống kê cho khoảng thời gian được chọn"""
        if self.start_date_picker.count() == 0 or self.end_date_picker.count() == 0:
            return
            
        start_date = datetime.strptime(self.start_date_picker.currentText(), '%d/%m/%Y').date()
        end_date = datetime.strptime(self.end_date_picker.currentText(), '%d/%m/%Y').date()
        
        if start_date > end_date:
            QMessageBox.warning(self, "Cảnh báo", "Ngày bắt đầu phải nhỏ hơn hoặc bằng ngày kết thúc!")
            return
        
        # Lọc dữ liệu theo khoảng thời gian
        date_range_data = [r for r in self.attendance_data 
                          if start_date <= datetime.fromisoformat(r['datetime']).date() <= end_date]
        
        # Tạo thống kê tổng quan
        total_days = (end_date - start_date).days + 1
        total_attendance = len(date_range_data)
        total_people = len(self.database)
        
        # Tính tỷ lệ điểm danh trung bình
        if total_days > 0 and total_people > 0:
            max_possible_attendance = total_days * total_people
            avg_attendance_rate = (total_attendance / max_possible_attendance) * 100
        else:
            avg_attendance_rate = 0
        
        # Cập nhật labels
        self.total_days_label.setText(str(total_days))
        self.total_attendance_label.setText(str(total_attendance))
        self.avg_attendance_rate_label.setText(f"{avg_attendance_rate:.1f}%")
        
        # Tạo thống kê theo ngày
        self.update_daily_stats_table(start_date, end_date, date_range_data)
        
        # Tạo thống kê theo người
        self.update_person_stats_table(start_date, end_date, date_range_data)
        
        QMessageBox.information(self, "Thành công", "Đã tạo thống kê thành công!")
    
    def update_daily_stats_table(self, start_date, end_date, date_range_data):
        """Cập nhật bảng thống kê theo ngày"""
        # Tạo danh sách tất cả các ngày trong khoảng
        all_dates = []
        current_date = start_date
        while current_date <= end_date:
            all_dates.append(current_date)
            current_date += timedelta(days=1)
        
        self.daily_stats_table.setRowCount(len(all_dates))
        
        for row, date in enumerate(all_dates):
            # Lọc dữ liệu theo ngày
            daily_data = [r for r in date_range_data 
                         if datetime.fromisoformat(r['datetime']).date() == date]
            
            # Ngày
            date_item = QTableWidgetItem(date.strftime('%d/%m/%Y'))
            self.daily_stats_table.setItem(row, 0, date_item)
            
            # Tổng số người
            total_people = len(self.database)
            total_item = QTableWidgetItem(str(total_people))
            self.daily_stats_table.setItem(row, 1, total_item)
            
            # Số người điểm danh
            present_count = len(daily_data)
            present_item = QTableWidgetItem(str(present_count))
            present_item.setBackground(QColor(200, 255, 200))
            self.daily_stats_table.setItem(row, 2, present_item)
            
            # Số người vắng
            absent_count = total_people - present_count
            absent_item = QTableWidgetItem(str(absent_count))
            absent_item.setBackground(QColor(255, 200, 200))
            self.daily_stats_table.setItem(row, 3, absent_item)
            
            # Tỷ lệ điểm danh
            attendance_rate = (present_count / total_people * 100) if total_people > 0 else 0
            rate_item = QTableWidgetItem(f"{attendance_rate:.1f}%")
            self.daily_stats_table.setItem(row, 4, rate_item)
    
    def update_person_stats_table(self, start_date, end_date, date_range_data):
        """Cập nhật bảng thống kê theo người"""
        total_people = len(self.database)
        self.person_stats_table.setRowCount(total_people)
        
        for row, person in enumerate(self.database.keys()):
            # Lọc dữ liệu theo người
            person_data = [r for r in date_range_data if r['name'] == person]
            
            # Tên
            name_item = QTableWidgetItem(person)
            self.person_stats_table.setItem(row, 0, name_item)
            
            # Số lần điểm danh
            attendance_count = len(person_data)
            count_item = QTableWidgetItem(str(attendance_count))
            self.person_stats_table.setItem(row, 1, count_item)
            
            # Tỷ lệ điểm danh
            total_days = (end_date - start_date).days + 1
            attendance_rate = (attendance_count / total_days * 100) if total_days > 0 else 0
            rate_item = QTableWidgetItem(f"{attendance_rate:.1f}%")
            self.person_stats_table.setItem(row, 2, rate_item)
            
            # Ngày điểm danh cuối cùng
            if person_data:
                last_attendance = max(person_data, key=lambda x: x['datetime'])
                last_date = datetime.fromisoformat(last_attendance['datetime']).strftime('%d/%m/%Y')
                last_item = QTableWidgetItem(last_date)
            else:
                last_item = QTableWidgetItem("Chưa điểm danh")
                last_item.setBackground(QColor(255, 200, 200))
            self.person_stats_table.setItem(row, 3, last_item)
    
    def export_statistics(self):
        """Xuất thống kê ra CSV"""
        try:
            # Tạo dữ liệu cho thống kê tổng quan
            overall_data = {
                'Chỉ số': ['Tổng số ngày', 'Tổng số lần điểm danh', 'Tỷ lệ điểm danh trung bình'],
                'Giá trị': [
                    self.total_days_label.text(),
                    self.total_attendance_label.text(),
                    self.avg_attendance_rate_label.text()
                ]
            }
            
            # Tạo dữ liệu cho thống kê theo ngày
            daily_data = []
            for row in range(self.daily_stats_table.rowCount()):
                daily_data.append({
                    'Ngày': self.daily_stats_table.item(row, 0).text(),
                    'Tổng Người': self.daily_stats_table.item(row, 1).text(),
                    'Điểm Danh': self.daily_stats_table.item(row, 2).text(),
                    'Vắng': self.daily_stats_table.item(row, 3).text(),
                    'Tỷ Lệ': self.daily_stats_table.item(row, 4).text()
                })
            
            # Tạo dữ liệu cho thống kê theo người
            person_data = []
            for row in range(self.person_stats_table.rowCount()):
                person_data.append({
                    'Tên': self.person_stats_table.item(row, 0).text(),
                    'Số Lần Điểm Danh': self.person_stats_table.item(row, 1).text(),
                    'Tỷ Lệ': self.person_stats_table.item(row, 2).text(),
                    'Ngày Cuối': self.person_stats_table.item(row, 3).text()
                })
            
            # Lưu thống kê vào CSV
            import pandas as pd
            current_date = datetime.now().strftime('%Y-%m-%d')
            
            # Lưu thống kê tổng quan
            overall_filename = os.path.join(self.attendance_csv_dir, f"statistics_overall_{current_date}.csv")
            pd.DataFrame(overall_data).to_csv(overall_filename, index=False, encoding='utf-8-sig', sep='\t')
            
            # Lưu thống kê theo ngày
            daily_filename = os.path.join(self.attendance_csv_dir, f"statistics_daily_{current_date}.csv")
            pd.DataFrame(daily_data).to_csv(daily_filename, index=False, encoding='utf-8-sig', sep='\t')
            
            # Lưu thống kê theo người
            person_filename = os.path.join(self.attendance_csv_dir, f"statistics_person_{current_date}.csv")
            pd.DataFrame(person_data).to_csv(person_filename, index=False, encoding='utf-8-sig', sep='\t')
            
            QMessageBox.information(self, "Thành công", f"Đã xuất thống kê ra CSV vào thư mục {self.attendance_csv_dir}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Không thể xuất thống kê CSV: {e}")

def main():
    app = QApplication(sys.argv)
    
    # Set application properties
    app.setApplicationName("Face Recognition Attendance System")
    app.setApplicationVersion("1.0")
    
    # Create and show main window
    window = AttendanceSystem()
    window.show()
    
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()
