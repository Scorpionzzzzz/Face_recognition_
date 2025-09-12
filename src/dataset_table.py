import os
import glob
import pandas as pd

def create_dataset_table():
    """Tạo bảng thống kê dataset dạng bảng"""
    
    print("🔍 ĐANG TẠO BẢNG THỐNG KÊ DATASET...")
    print("=" * 60)
    
    # Thống kê thư mục FRAME_DATASET
    frame_dir = "FRAME_DATASET"
    if not os.path.exists(frame_dir):
        print("❌ Không tìm thấy thư mục FRAME_DATASET!")
        return
    
    # Thu thập dữ liệu
    splits = ['train', 'val', 'test']
    people_data = {}
    
    for split in splits:
        split_path = os.path.join(frame_dir, split)
        if os.path.exists(split_path):
            print(f"📂 Đang xử lý {split.upper()}...")
            
            # Lấy danh sách người
            people = [d for d in os.listdir(split_path) if os.path.isdir(os.path.join(split_path, d))]
            
            for person in people:
                person_path = os.path.join(split_path, person)
                images = glob.glob(os.path.join(person_path, "*.jpg"))
                image_count = len(images)
                
                if person not in people_data:
                    people_data[person] = {}
                
                people_data[person][split] = image_count
    
    # Tạo DataFrame
    print("\n📊 ĐANG TẠO BẢNG...")
    
    # Chuyển đổi dữ liệu thành DataFrame
    table_data = []
    for person in sorted(people_data.keys()):
        row = {
            'Tên': person,
            'Train': people_data[person].get('train', 0),
            'Val': people_data[person].get('val', 0),
            'Test': people_data[person].get('test', 0)
        }
        # Tính tổng
        row['Tổng'] = row['Train'] + row['Val'] + row['Test']
        table_data.append(row)
    
    # Tạo DataFrame
    df = pd.DataFrame(table_data)
    
    # Sắp xếp theo tổng số ảnh (giảm dần)
    df = df.sort_values('Tổng', ascending=False)
    
    # Hiển thị bảng
    print("\n" + "=" * 80)
    print("📊 BẢNG THỐNG KÊ DATASET")
    print("=" * 80)
    
    # Tạo header
    header = f"{'Tên':<25} {'Train':<8} {'Val':<8} {'Test':<8} {'Tổng':<8}"
    print(header)
    print("-" * 80)
    
    # Hiển thị từng dòng
    for _, row in df.iterrows():
        line = f"{row['Tên']:<25} {row['Train']:<8} {row['Val']:<8} {row['Test']:<8} {row['Tổng']:<8}"
        print(line)
    
    # Hiển thị tổng
    print("-" * 80)
    total_row = {
        'Tên': 'TỔNG CỘNG',
        'Train': df['Train'].sum(),
        'Val': df['Val'].sum(),
        'Test': df['Test'].sum(),
        'Tổng': df['Tổng'].sum()
    }
    total_line = f"{total_row['Tên']:<25} {total_row['Train']:<8} {total_row['Val']:<8} {total_row['Test']:<8} {total_row['Tổng']:<8}"
    print(total_line)
    
    # Thống kê tổng quan
    print("\n" + "=" * 60)
    print("📈 THỐNG KÊ TỔNG QUAN:")
    print("-" * 30)
    print(f"👥 Tổng số người: {len(df)}")
    print(f"🖼️ Tổng số ảnh: {total_row['Tổng']}")
    print(f"📊 Trung bình ảnh/người: {total_row['Tổng']/len(df):.1f}")
    
    # Lưu bảng ra file CSV
    output_file = "dataset_statistics_table.csv"
    df.to_csv(output_file, index=False, encoding='utf-8-sig', sep='\t')
    print(f"\n💾 Đã lưu bảng thống kê vào: {output_file}")
    
    # Lưu bảng ra file Excel (nếu có pandas và openpyxl)
    try:
        excel_file = "dataset_statistics_table.xlsx"
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Dataset Statistics', index=False)
            
            # Lấy worksheet để định dạng
            worksheet = writer.sheets['Dataset Statistics']
            
            # Định dạng header
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Style cho header
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Border style
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Áp dụng style cho header
            for col in range(1, 6):  # A đến E
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Áp dụng style cho các dòng dữ liệu
            for row in range(2, len(df) + 2):
                for col in range(1, 6):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = thin_border
                    
                    # Căn giữa cho các cột số
                    if col > 1:
                        cell.alignment = Alignment(horizontal="center")
            
            # Điều chỉnh độ rộng cột
            worksheet.column_dimensions['A'].width = 25  # Tên
            worksheet.column_dimensions['B'].width = 10  # Train
            worksheet.column_dimensions['C'].width = 10  # Val
            worksheet.column_dimensions['D'].width = 10  # Test
            worksheet.column_dimensions['E'].width = 10  # Tổng
        
        print(f"✅ Đã lưu bảng thống kê Excel: {excel_file}")
        
    except Exception as e:
        print(f"⚠️ Không thể lưu Excel (cần pandas và openpyxl): {e}")
    
    return df

def main():
    """Hàm chính"""
    try:
        df = create_dataset_table()
        print(f"\n🎉 HOÀN THÀNH! Bảng thống kê có {len(df)} người")
        
    except Exception as e:
        print(f"❌ Lỗi khi tạo bảng thống kê: {e}")

if __name__ == "__main__":
    main()
